from openpyxl import load_workbook


def 保存聊天会话df(output_fpath, df, template_fpath='tpl3.xlsx', sheet_name='会话表'):
    # 加载模板
    print('使用模版文件：', template_fpath)
    template = load_workbook(template_fpath, rich_text=True)
    sheet = template[sheet_name]
    
    for row_idx, row_data in enumerate(df.values, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value.strip())

    
        
    # 删除工作表
    template.remove(template['提示词'])

    template.save(output_fpath)
